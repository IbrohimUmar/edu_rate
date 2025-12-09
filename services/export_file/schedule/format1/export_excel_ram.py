from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.formatting.rule import DataBarRule
import io

from models.models.export_file import ExportFile
from models.models.survey import SurveyQuestion, SurveyAnswerOption


def export_file_schedule_format1_excel_ram(data, survey, export_file=None):
    # ======================================
    # Sabit header
    # ======================================
    fixed_headers = ["Sana", "Fakultet", "Guruh", "Fan", "O‘qituvchi"]
    # ======================================
    # 1) Workbook Oluştur veya Mevcut Aç
    # ======================================
    if export_file and export_file.file:
        export_file.file.open('rb')
        wb = load_workbook(filename=export_file.file)
        ws = wb.active
        row_idx = ws.max_row + 1  # Mevcut satırdan devam

        # Mevcut workbookta question_answer_map ve answer_cols çıkarmak
        questions = SurveyQuestion.objects.filter(survey=survey).order_by("order_position")
        question_answer_map = {}
        answer_cols = []
        answer_types = []
        col = len(fixed_headers) + 1
        for q in questions:
            opts = SurveyAnswerOption.objects.filter(question=q).order_by("type")
            question_answer_map[q] = list(opts)
            for opt in opts:
                answer_cols.append(get_column_letter(col))
                answer_types.append(opt.type)
                col += 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule KPI Dynamic"
        row_idx = 3  # headerler için hazır

        # ======================================
        # HEADERLER
        # ======================================
        questions = SurveyQuestion.objects.filter(survey=survey).order_by("order_position")
        question_answer_map = {}
        for q in questions:
            opts = SurveyAnswerOption.objects.filter(question=q).order_by("type")
            question_answer_map[q] = list(opts)

        # Fixed Headers
        col = 1
        for head in fixed_headers:
            col_letter = get_column_letter(col)
            ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
            ws.cell(row=1, column=col, value=head)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.column_dimensions[col_letter].width = 25
            col += 1

        # Dynamic Headers
        for q, opts in question_answer_map.items():
            span = len(opts)
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
            ws.cell(row=1, column=col, value=q.name)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=1, column=col).font = Font(bold=True)
            col += span

        col = len(fixed_headers) + 1
        answer_cols = []
        answer_types = []
        for q, opts in question_answer_map.items():
            for opt in opts:
                cell = ws.cell(row=2, column=col, value=opt.name)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, textRotation=90)
                ws.column_dimensions[get_column_letter(col)].width = max(len(opt.name) // 2, 6)
                answer_cols.append(get_column_letter(col))
                answer_types.append(opt.type)
                col += 1
    # ======================================
    # 2) DATA EKLEME
    # ======================================
    for item in data:
        schedule = item["schedule"]
        qa_list = item["questions_and_answer_count"]

        # Fixed Fields
        values = [
            schedule.lesson_date.strftime("%Y-%m-%d"),
            schedule.faculty.name if schedule.faculty else "",
            schedule.group.name if schedule.group else "",
            schedule.subject.name if schedule.subject else "",
            f"{schedule.employee.full_name}",
        ]
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

        # Dynamic Answers
        col = len(fixed_headers) + 1
        for q, opts in question_answer_map.items():
            found = next((x for x in qa_list if x["question_id"] == q.id), None)
            for opt in opts:
                count_val = 0
                if found:
                    matched = next((a for a in found["answers"] if a["id"] == opt.id), None)
                    if matched:
                        count_val = matched["send_count"]
                ws.cell(row=row_idx, column=col, value=count_val)
                col += 1
        row_idx += 1

    # ======================================
    # 3) BORDER & STYLE
    # ======================================
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in r:
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ======================================
    # 4) CONDITIONAL FORMATTING
    # ======================================
    color_map = {"1": "70AD47", "2": "ED7D31", "3": "C00000"}
    for col_letter, type_choice in zip(answer_cols, answer_types):
        color = color_map.get(type_choice, "4472C4")
        databar = DataBarRule(start_type="min", end_type="max", color=color)
        ws.conditional_formatting.add(f"{col_letter}3:{col_letter}{ws.max_row}", databar)

    # ======================================
    # 5) EXCEL RAMDE KAYDET
    # ======================================
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # ExportFile objesi varsa güncelle, yoksa oluştur
    if not export_file:
        export_file = ExportFile.objects.create(user=survey.user, name="dynamic_survey.xlsx")

    export_file.file.save("dynamic_survey.xlsx", content=io.BytesIO(output.getvalue()), save=True)
    # export_file.status = "done"
    export_file.save()

    return export_file
