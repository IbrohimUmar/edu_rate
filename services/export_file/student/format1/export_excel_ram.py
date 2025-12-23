from openpyxl import Workbook, load_workbook
from django.core.files.base import ContentFile
from io import BytesIO

from models.models.export_file import ExportFile


def export_students_excel_ram(data_rows, export_file: ExportFile):
    if export_file.file:
        # Dosya varsa aç ve append et
        wb = load_workbook(export_file.file.path)
        ws = wb.active
    else:
        # İlk sayfa
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"

        # Header ekle
        ws.append([
            "Hemis ID",
            "F.I.Sh",
            "Fakultet",
            "Guruh",
            "Ro'yxatdan o'tganmi"
        ])

    for row in data_rows:
        ws.append(row)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    export_file.file.save(
        export_file.name + ".xlsx",
        ContentFile(buffer.read()),
        save=False
    )
