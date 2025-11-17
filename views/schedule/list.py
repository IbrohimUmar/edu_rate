import io, datetime
import random
import tempfile

from django.db.models import Q, Count, Sum, F, OuterRef, Subquery, Case, Value, FloatField, When, IntegerField
from django.http import HttpResponse, FileResponse
from django.shortcuts import render, redirect
from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from openpyxl.styles import Alignment, Font, Side, Border
from openpyxl.worksheet.cell_range import CellRange

from django.core.paginator import Paginator
from django.utils.timezone import now

from django.contrib.auth.decorators import login_required
from django.shortcuts import render

from models.models.answer import Answer, AnswerDetail
from models.models.schedule import Schedule
from django.db.models import Count, Q

from models.models.survey import SurveyAnswerOption, SurveyQuestion, Survey


# from models.models.schedule_point import SchedulePoint


@login_required(login_url='login')
def schedule_list(request):
    search_query = request.GET.get('search', '').strip()
    export_to_excel = request.GET.get('export') == 'excel'
    date_str = request.GET.get('date')
    try:
        selected_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else now().date()
    except ValueError:
        selected_date = now().date()  # noto‘g‘ri format bo‘lsa, bugungi kun
    schedules = Schedule.objects.filter(
                                        lesson_date__date=selected_date
                                        ).order_by('lesson_date')
    schedules = (
        schedules
        .annotate(
            group_student_count_qs=Count("group__student", distinct=True),
            answer_send_count_qs=Count(
                "answer",
                filter=Q(answer__answer_submitted_at__isnull=False),  # sadece cevap gönderenler
                distinct=True
            )
        )
    )
    send_answer_type = request.GET.get("send_answer", '0')
    if send_answer_type != '0':
        if send_answer_type == '1':
            schedules = schedules.filter(answer_send_count_qs__gt=0)
        else:
            schedules = schedules.exclude(answer_send_count_qs__gt=0)
    if search_query:
        schedules = schedules.filter(
            Q(employee__first_name__icontains=search_query) |
            Q(employee__second_name__icontains=search_query) |
            Q(employee__third_name__icontains=search_query)
        )


    data = []
    for s in schedules:
        questions = []
        survey_questions = SurveyQuestion.objects.filter(survey=s.survey).order_by("order_position")
        for q in survey_questions:
            question_answer_count = []
            survey_question_options = SurveyAnswerOption.objects.filter(question=q)
            for o in survey_question_options:
                count = AnswerDetail.objects.filter(survey_question=q, answer__schedule=s.id, survey_answer_option=o).count()
                question_answer_count.append({
                    'id':o.id,'name':o.name, 'type':o.type,
                    'send_count':count
                })
            questions.append(
                {"question_id":q.id, 'answers':question_answer_count}
            )
        data.append(
            {
                'schedule_id': s.id,
                'schedule': s,
                'questions_and_answer_count': questions
            }
        )


    # print(data)

    # for d in data:
    #     print(d)
    #     print('\n')


    if export_to_excel:
        schedules = schedules.filter(survey__isnull=False)
        survey = Survey.objects.get(id=6)
        print(schedules)
        print('ishladi')
        # return export_schedule_summary_excel(schedules, survey)
        # wb = export_schedule_summary_excel(schedules, survey)
        # wb = export_schedule_summary_excel(schedules, survey)
        # return excel_download(wb, "dars_baholash.xlsx")
        return export_schedule_to_excel(data, survey)
    page = request.GET.get('page', 1)
    paginator = Paginator(schedules, 50)  # sahifada 25 ta qator
    queryset = paginator.get_page(page)  # ← SENING HTML'DAGI NOM
    queryset_count = paginator.count     # ← umumiy ma'lumotlar soni
    context = {
        'queryset': queryset,
        'queryset_count': queryset_count,
        'selected_date': selected_date
    }
    return render(request, "schedule/list.html", context)



# def export_schedule_to_excel(data, survey):
#     from openpyxl import Workbook
#     from openpyxl.styles import Alignment, Font, Border, Side
#
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Schedule KPI Dynamic"
#
#     # ======================================
#     # 1) SAVOLLARNI DYNAMIC YUKLASH
#     # ======================================
#     questions = SurveyQuestion.objects.filter(survey=survey).order_by("order_position")
#
#     question_answer_map = {}
#     total_dynamic_cols = 0
#
#     for q in questions:
#         opts = SurveyAnswerOption.objects.filter(question=q).order_by("type")
#         question_answer_map[q] = list(opts)
#         total_dynamic_cols += len(opts)
#
#     # ======================================
#     # 2) HEADER — 1-QATOR (Savol nomlari merge)
#     # ======================================
#
#     ws.cell(row=1, column=1, value="Schedule ID")
#     ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
#     ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
#     ws.cell(row=1, column=1).font = Font(bold=True)
#
#     col = 2
#     for q, opts in question_answer_map.items():
#         span = len(opts)
#         ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
#         ws.cell(row=1, column=col, value=q.name)
#         ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
#         ws.cell(row=1, column=col).font = Font(bold=True)
#         col += span
#
#     # ======================================
#     # 3) HEADER — 2-QATOR (Javob nomlari)
#     # ======================================
#
#     col = 2
#     for q, opts in question_answer_map.items():
#         for opt in opts:
#             ws.cell(row=2, column=col, value=opt.name)
#             ws.cell(row=2, column=col).alignment = Alignment(horizontal="center", vertical="center")
#             col += 1
#
#     # ======================================
#     # 4) MA'LUMOTLARNI YUKLASH
#     # ======================================
#
#     row_idx = 3
#     for item in data:   # sening keltirgan massiv
#         schedule_id = item["schedule_id"]
#         qa_list = item["questions_and_answer_count"]
#
#         ws.cell(row=row_idx, column=1, value=schedule_id)
#
#         col = 2
#
#         # har bir savol boyicha send_count joylashtiramiz
#         for q, opts in question_answer_map.items():
#
#             # shu schedule uchun aynan shu savolni topish
#             found = next((x for x in qa_list if x["question_id"] == q.id), None)
#
#             if not found:
#                 # Bu schedule'ga bu savolga javob yo'q
#                 for _ in opts:
#                     ws.cell(row=row_idx, column=col, value=0)
#                     col += 1
#                 continue
#
#             # Cevaplarni joylashtirish
#             answer_list = found["answers"]
#
#             for opt in opts:
#                 matched = next((a for a in answer_list if a["id"] == opt.id), None)
#                 ws.cell(
#                     row=row_idx,
#                     column=col,
#                     value=matched["send_count"] if matched else 0
#                 )
#                 col += 1
#
#         row_idx += 1
#
#     # ======================================
#     # 5) STYLE – BORDER, ALIGN
#     # ======================================
#
#     thin = Side(border_style="thin", color="000000")
#     border = Border(left=thin, right=thin, top=thin, bottom=thin)
#
#     max_row = ws.max_row
#     max_col = ws.max_column
#
#     for r in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
#         for c in r:
#             c.border = border
#             c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#
#     # ======================================
#     # 6) EXPORT
#     # ======================================
#
#     output = io.BytesIO()
#     wb.save(output)
#     output.seek(0)
#
#     filename = "dynamic_survey.xlsx"
#     response = HttpResponse(
#         output,
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
#     response['Content-Disposition'] = f'attachment; filename="{filename}"'
#
#     return response

def export_schedule_to_excel_old(data, survey):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule KPI Dynamic"

    # ======================================
    # 1) SAVOLLARNI YUKLASH
    # ======================================
    questions = SurveyQuestion.objects.filter(survey=survey).order_by("order_position")

    question_answer_map = {}
    for q in questions:
        opts = SurveyAnswerOption.objects.filter(question=q).order_by("type")
        question_answer_map[q] = list(opts)

    # ======================================
    # 2) HEADER — Fixed 5 Column
    # ======================================

    fixed_headers = ["Sana", "Fakultet", "Guruh", "Fan", "O‘qituvchi"]

    col = 1
    for head in fixed_headers:
        col_letter = get_column_letter(col)

        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        ws.cell(row=1, column=col, value=head)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.column_dimensions[col_letter].width = 25
        col += 1

    # ======================================
    # 3) HEADER — Dynamic Questions
    # ======================================

    for q, opts in question_answer_map.items():
        span = len(opts)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
        ws.cell(row=1, column=col, value=q.name)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=1, column=col).font = Font(bold=True)
        col += span

    # ======================================
    # 4) SUB HEADERS (Answer Options)
    # ======================================
    col = len(fixed_headers) + 1

    for q, opts in question_answer_map.items():
        for opt in opts:
            cell = ws.cell(row=2, column=col, value=opt.name)

            # 90° AYLANISH
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
                textRotation=90
            )
            # Ustun width (90° bo'lgani uchun ENGA emas BALANDLIKKА o‘ynaydi)
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 6  # 6–8 ideal

            col += 1
    # ======================================
    # 5) ROW DATAS
    # ======================================

    row_idx = 3
    for item in data:

        schedule = item["schedule"]     # <-- sen dictionaryda schedule objectni ham berayapsan
        qa_list = item["questions_and_answer_count"]

        # --- FIXED FIELDS ---
        values = [
            schedule.lesson_date.strftime("%Y-%m-%d"),
            schedule.faculty.name if schedule.faculty else "",
            schedule.group.name if schedule.group else "",
            schedule.subject.name if schedule.subject else "",
            f"{schedule.employee.full_name}",
        ]

        # avval fixed kolonlarni qo‘yamiz
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

        # --- DYNAMIC ANSWER COUNTS ---
        col = len(fixed_headers) + 1

        for q, opts in question_answer_map.items():

            found = next((x for x in qa_list if x["question_id"] == q.id), None)

            for opt in opts:
                if found:
                    matched = next((a for a in found["answers"] if a["id"] == opt.id), None)
                    ws.cell(row=row_idx, column=col, value=matched["send_count"] if matched else 0)
                else:
                    ws.cell(row=row_idx, column=col, value=0)
                col += 1

        row_idx += 1

    # ======================================
    # 6) STYLE
    # ======================================

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column

    for r in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for c in r:
            # rotation berilganlarni buzmaymiz
            if c.alignment and c.alignment.textRotation:
                # faqat border qo'yamiz
                c.border = border
            else:
                # rotation yo'q bo'lsa normal alignment
                c.border = border
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # ======================================
    # 7) EXPORT
    # ======================================

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="dynamic_survey.xlsx"'

    return response



from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter
import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side

from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter
import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side

def export_schedule_to_excel(data, survey):
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule KPI Dynamic"

    # ======================================
    # 1) SAVOLLARNI YUKLASH
    # ======================================
    questions = SurveyQuestion.objects.filter(survey=survey).order_by("order_position")

    question_answer_map = {}
    for q in questions:
        opts = SurveyAnswerOption.objects.filter(question=q).order_by("type")
        question_answer_map[q] = list(opts)

    # ======================================
    # 2) HEADER — Fixed 5 Column
    # ======================================
    fixed_headers = ["Sana", "Fakultet", "Guruh", "Fan", "O‘qituvchi"]

    col = 1
    for head in fixed_headers:
        col_letter = get_column_letter(col)
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        ws.cell(row=1, column=col, value=head)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.column_dimensions[col_letter].width = 25
        col += 1

    # ======================================
    # 3) HEADER — Dynamic Questions
    # ======================================
    for q, opts in question_answer_map.items():
        span = len(opts)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
        ws.cell(row=1, column=col, value=q.name)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=1, column=col).font = Font(bold=True)
        col += span

    # ======================================
    # 4) SUB HEADERS (Answer Options)
    # ======================================
    col = len(fixed_headers) + 1
    answer_cols = []  # progress bar uchun
    answer_types = []  # har bir ustunning type_choice

    for q, opts in question_answer_map.items():
        for opt in opts:
            cell = ws.cell(row=2, column=col, value=opt.name)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, textRotation=90)

            # Ustun kengligini savol nomi uzunligiga qarab oshiramiz
            ws.column_dimensions[get_column_letter(col)].width = max(len(opt.name) // 2, 6)
            answer_cols.append(get_column_letter(col))
            answer_types.append(opt.type)
            col += 1
    # ======================================
    # 5) ROW DATAS
    # ======================================
    row_idx = 3
    for item in data:
        schedule = item["schedule"]
        qa_list = item["questions_and_answer_count"]

        # --- FIXED FIELDS ---
        values = [
            schedule.lesson_date.strftime("%Y-%m-%d"),
            schedule.faculty.name if schedule.faculty else "",
            schedule.group.name if schedule.group else "",
            schedule.subject.name if schedule.subject else "",
            f"{schedule.employee.full_name}",
        ]

        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

        # --- DYNAMIC ANSWER COUNTS ---
        col = len(fixed_headers) + 1

        for q, opts in question_answer_map.items():
            found = next((x for x in qa_list if x["question_id"] == q.id), None)
            for opt in opts:
                count_val = 0
                if found:
                    matched = next((a for a in found["answers"] if a["id"] == opt.id), None)
                    if matched:
                        count_val = matched["send_count"]
                ws.cell(row=row_idx, column=col, value=count_val)
                col += 1
        row_idx += 1

    # ======================================
    # 6) STYLE & BORDER
    # ======================================
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column

    for r in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for c in r:
            if c.alignment and c.alignment.textRotation:
                c.border = border
            else:
                c.border = border
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ======================================
    # 7) CONDITIONAL FORMATTING (Type bo'yicha progress bar)
    # ======================================
    color_map = {
        "1": "70AD47",  # Pozitiv → green
        "2": "ED7D31",  # Normal → orange
        "3": "C00000",  # Negative → red
    }

    for col_letter, type_choice in zip(answer_cols, answer_types):
        color = color_map.get(type_choice, "4472C4")  # default blue
        databar = DataBarRule(start_type="min", end_type="max", color=color)
        ws.conditional_formatting.add(f"{col_letter}3:{col_letter}{max_row}", databar)

    # ======================================
    # 8) EXPORT
    # ======================================
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="dynamic_survey.xlsx"'

    return response



#
# def export_schedule_to_excel(schedules):
#     pass
#     schedules_id = list(schedules.values_list('id', flat=True))
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Schedule KPI"
#
#     # --- ÜST BAŞLIKLAR ---
#     headers_top = [
#         ("Sana", 1),
#         ("Fakultet", 1),
#         ("Guruh", 1),
#         ("Fan", 1),
#         ("O'qituvchi", 1),
#         ("Pedagog darsga keldimi", 3),
#         ("O'qituvchining nutq va muomala madaniyati", 4),
#         ("O'tilayotgan mavzuning amaliyotga bog'langanligi", 3),
#         ("Yakunlangan darsga shaxsiy qoldirgan bahosi", 4),
#     ]
#
#     col = 1
#     for title, span in headers_top:
#         start_row = 1
#         end_row = 2 if span == 1 else 1
#         start_col = col
#         end_col = col + span - 1
#         ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
#         ws.cell(row=1, column=col, value=title)
#         ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#         ws.cell(row=1, column=col).font = Font(bold=True)
#         col += span
#
#     # --- ALT BAŞLIKLAR ---
#     subheaders = [
#         "Ha", "Yo'q", "Javobsiz",
#         "A'lo", "Qoniqarli", "Qoniqarsiz", "Javobsiz",
#         "Ha", "Yo'q", "Javobsiz",
#         "A'lo", "Qoniqarli", "Qoniqarsiz", "Javobsiz",
#     ]
#     start_sub_col = 6
#     for i, header in enumerate(subheaders, start=start_sub_col):
#         ws.cell(row=2, column=i, value=header)
#         ws.cell(row=2, column=i).alignment = Alignment(textRotation=90)
#
#     for i in range(1, 6):
#         ws.cell(row=1, column=i).alignment = Alignment(horizontal="center", vertical="center")
#         ws.cell(row=1, column=i).font = Font(bold=True)
#
#     # --- VERİLER ---
#     qs = (
#         SchedulePoint.objects.filter(schedule_id__in=schedules_id)
#         .values(
#             "schedule_id",
#             "schedule__lesson_date",
#             "schedule__group__name",
#             "schedule__subject__name",
#             "schedule__faculty__name",
#             "schedule__employee__first_name",
#             "schedule__employee__second_name",
#         )
#         .annotate(
#             teacher_present_0=Count(Case(When(is_teacher_present="0", then=1), output_field=IntegerField())),
#             teacher_present_1=Count(Case(When(is_teacher_present="1", then=1), output_field=IntegerField())),
#             teacher_present_2=Count(Case(When(is_teacher_present="2", then=1), output_field=IntegerField())),
#             speech_0=Count(Case(When(teacher_speech_and_culture="0", then=1), output_field=IntegerField())),
#             speech_1=Count(Case(When(teacher_speech_and_culture="1", then=1), output_field=IntegerField())),
#             speech_2=Count(Case(When(teacher_speech_and_culture="2", then=1), output_field=IntegerField())),
#             speech_3=Count(Case(When(teacher_speech_and_culture="3", then=1), output_field=IntegerField())),
#             relevance_0=Count(Case(When(topic_practical_relevance="0", then=1), output_field=IntegerField())),
#             relevance_1=Count(Case(When(topic_practical_relevance="1", then=1), output_field=IntegerField())),
#             relevance_2=Count(Case(When(topic_practical_relevance="2", then=1), output_field=IntegerField())),
#             feedback_0=Count(Case(When(lesson_feedback="0", then=1), output_field=IntegerField())),
#             feedback_1=Count(Case(When(lesson_feedback="1", then=1), output_field=IntegerField())),
#             feedback_2=Count(Case(When(lesson_feedback="2", then=1), output_field=IntegerField())),
#             feedback_3=Count(Case(When(lesson_feedback="3", then=1), output_field=IntegerField())),
#         )
#     )
#
#     for row in qs:
#         date_str = row["schedule__lesson_date"].strftime("%Y-%m-%d") if row["schedule__lesson_date"] else ""
#         ws.append([
#             date_str,
#             row["schedule__faculty__name"],
#             row["schedule__group__name"],
#             row["schedule__subject__name"],
#             f"{row['schedule__employee__first_name']} {row['schedule__employee__second_name']}",
#             row["teacher_present_1"], row["teacher_present_2"], row["teacher_present_0"],
#             row["speech_3"], row["speech_1"], row["speech_2"], row["speech_0"],
#             row["relevance_1"], row["relevance_2"], row["relevance_0"],
#             row["feedback_3"], row["feedback_1"], row["feedback_2"], row["feedback_0"],
#         ])
#
#     # --- STİL ---
#     thin = Side(border_style="thin", color="000000")
#     border = Border(left=thin, right=thin, top=thin, bottom=thin)
#     max_row = ws.max_row
#     max_col = ws.max_column
#
#     for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
#         for cell in row:
#             align = cell.alignment
#             cell.alignment = Alignment(
#                 horizontal="center", vertical="center", wrap_text=True, textRotation=align.textRotation
#             )
#             cell.border = border
#
#     ws.row_dimensions[1].height = 74
#     ws.row_dimensions[2].height = 60
#
#     for c in range(1, 6):
#         ws.column_dimensions[chr(64 + c)].width = 18
#     for c in range(6, max_col + 1):
#         ws.column_dimensions[chr(64 + c)].width = 6
#
#     # --- DATA BAR ---
#     databar_blue = DataBarRule(start_type="min", end_type="max", color="4472C4")
#     databar_green = DataBarRule(start_type="min", end_type="max", color="70AD47")
#     databar_orange = DataBarRule(start_type="min", end_type="max", color="ED7D31")
#
#     ws.conditional_formatting.add(f"F3:F{max_row}", databar_green)
#     ws.conditional_formatting.add(f"I3:I{max_row}", databar_green)
#     ws.conditional_formatting.add(f"M3:M{max_row}", databar_green)
#     ws.conditional_formatting.add(f"P3:P{max_row}", databar_green)
#
#     ws.conditional_formatting.add(f"G3:G{max_row}", databar_orange)
#     ws.conditional_formatting.add(f"K3:K{max_row}", databar_orange)
#     ws.conditional_formatting.add(f"N3:N{max_row}", databar_orange)
#     ws.conditional_formatting.add(f"R3:R{max_row}", databar_orange)
#
#     ws.conditional_formatting.add(f"H3:H{max_row}", databar_blue)
#     ws.conditional_formatting.add(f"L3:L{max_row}", databar_blue)
#     ws.conditional_formatting.add(f"O3:O{max_row}", databar_blue)
#     ws.conditional_formatting.add(f"S3:S{max_row}", databar_blue)
#
#     # --- ÇIKTI ---
#     output = io.BytesIO()
#     wb.save(output)
#     output.seek(0)
#
#     now_str = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
#     filename = f'fan_boyicha_statistika_{now_str}.xlsx'
#     response = HttpResponse(
#         output,
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = f'attachment; filename="{filename}"'
#     return response
#


import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# modelleri kendi app'ine göre değiştir

def safe_name(obj, attr="name", default=""):
    if obj is None:
        return default
    return getattr(obj, attr, default) or default

def export_schedule_summary_excel(schedules, survey=None):
    """
    schedules: Schedule queryset veya iterable
    survey: opsiyonel Survey objesi (eğer None ise schedules içinden alınmaya çalışılır)
    Döndürür: openpyxl.Workbook objesi
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Dars baholash"

    bold_center = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Eğer survey verilmemişse schedules'tan almayı dene (ilk schedule içindeki survey)
    if survey is None:
        try:
            sample = next(iter(schedules))
            survey = getattr(sample, "survey", None)
        except StopIteration:
            survey = None

    if survey is None:
        raise ValueError("Survey parametresi boş. Ya survey ver, ya schedules içinde en az 1 schedule olmalı.")

    # ===== 1) Sorular ve seçenekler =====
    questions = list(SurveyQuestion.objects.filter(survey=survey).order_by("order_position"))
    # question.id -> list(opt objs)
    question_option_map = {}
    for q in questions:
        opts = list(SurveyAnswerOption.objects.filter(question=q).order_by("id"))
        question_option_map[q.id] = opts

    # ===== 2) Header =====
    fixed_headers = ["Sana", "Fakultet", "Guruh", "Fan", "O‘qituvchi"]
    start_dynamic_col = 6  # dinamik sütunlar 6. sütundan başlıyor
    col = 1

    for h in fixed_headers:
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold_center
        cell.alignment = center
        cell.border = thin_border
        col += 1

    # Dinamik soru + seçenek sütunları
    # Eğer questions boşsa col_index hiç artmaz; ama start_dynamic_col sabit kalsın
    for q in questions:
        opts = question_option_map.get(q.id, [])
        for opt in opts:
            val = f"{q.name}\n({opt.name})"
            cell = ws.cell(row=1, column=col, value=val)
            cell.alignment = center
            cell.font = bold_center
            cell.border = thin_border
            col += 1

    # ===== 3) Veri doldurma =====
    row = 2

    schedule_list = list(schedules)
    schedule_ids = [s.id for s in schedule_list]

    # Eğer hiç schedule yoksa yine boş workbook dönecek (header var)
    if schedule_ids:
        answers_qs = (
            Answer.objects
            .filter(schedule_id__in=schedule_ids)
            .prefetch_related("answerdetail_set")
        )
    else:
        answers_qs = []

    # Map: schedule_id -> list of answers
    answers_by_schedule = {}
    for ans in answers_qs:
        answers_by_schedule.setdefault(ans.schedule_id, []).append(ans)

    # col_index'i garantiye al (başlangıç dinamik kolonu)
    # Eğer hiç soru yoksa column_count = start_dynamic_col - 1
    col_index = start_dynamic_col

    for sch in schedule_list:
        # init stats: question_id -> opt_id -> count
        stats = {q.id: {opt.id: 0 for opt in question_option_map.get(q.id, [])} for q in questions}

        answers = answers_by_schedule.get(sch.id, [])

        for ans in answers:
            # answerdetail_set zaten prefetch edilmiş
            for detail in getattr(ans, "answerdetail_set").all():
                qid = getattr(detail, "survey_question_id", None)
                oid = getattr(detail, "survey_answer_option_id", None)
                if qid and oid and qid in stats and oid in stats[qid]:
                    stats[qid][oid] += 1

        # Satırı yaz
        lesson_date = getattr(sch, "lesson_date", None)
        date_str = lesson_date.strftime("%d.%m.%Y") if lesson_date else ""
        ws.cell(row=row, column=1, value=date_str)
        ws.cell(row=row, column=2, value=safe_name(getattr(sch, "faculty", None)))
        ws.cell(row=row, column=3, value=safe_name(getattr(sch, "group", None)))
        ws.cell(row=row, column=4, value=safe_name(getattr(sch, "subject", None)))
        emp = getattr(sch, "employee", None)
        # employee için önce full_name, sonra name
        emp_name = getattr(emp, "full_name", None) or getattr(emp, "name", None) or ""
        ws.cell(row=row, column=5, value=emp_name)

        # stil uygula
        for i in range(1, 6):
            c = ws.cell(row=row, column=i)
            c.alignment = center
            c.border = thin_border

        # dinamik sutunlar
        col_index = start_dynamic_col
        for q in questions:
            for opt in question_option_map.get(q.id, []):
                val = stats[q.id].get(opt.id, 0)
                c = ws.cell(row=row, column=col_index, value=val)
                c.alignment = center
                c.border = thin_border
                col_index += 1

        row += 1

    # column genişlikleri (hesaplanmış son col_index'i kullan)
    # Eğer hiç soru yoksa col_index start_dynamic_col'da kalır -> genişlikleri en az o kadar ayarlar
    max_col = col_index if 'col_index' in locals() else start_dynamic_col
    for c_idx in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 18

    return wb


def excel_download(workbook, filename="report.xlsx"):
    import io
    from django.http import HttpResponse

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response